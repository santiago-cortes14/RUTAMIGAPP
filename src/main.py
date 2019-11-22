from flask import Flask, render_template, request
from flask_wtf import FlaskForm
from wtforms import StringField,PasswordField
from wtforms.validators import InputRequired, Length,AnyOf
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
import pandas as pd
import numpy as np
import time
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import xlsxwriter
import openpyxl
import serial

def writezero():
    worksheet.write('A1', 'Codigos')
    worksheet.write(1,0,1001)
    worksheet.write(2,0,1002)
    worksheet.write(3,0,1003)
    worksheet.write(4,0,1004)
    worksheet.write(5,0,1005)
    worksheet.write(6,0,1006)
    worksheet.write(7,0,1007)
    worksheet.write(8,0,1008)
    worksheet.write(9,0,1009)
    worksheet.write(10,0,1010)
    worksheet.write('B1', 'Centinela')
    worksheet.write(1,1,0)
    worksheet.write(2,1,0)
    worksheet.write(3,1,0)
    worksheet.write(4,1,0)
    worksheet.write(5,1,0)
    worksheet.write(6,1,0)
    worksheet.write(7,1,0)
    worksheet.write(8,1,0)
    worksheet.write(9,1,0)
    worksheet.write(10,1,0)
    worksheet.write('C1', 'Centifinal')
    worksheet.write(1,2,0)
    worksheet.write(2,2,0)
    worksheet.write(3,2,0)
    worksheet.write(4,2,0)
    worksheet.write(5,2,0)
    worksheet.write(6,2,0)
    worksheet.write(7,2,0)
    worksheet.write(8,2,0)
    worksheet.write(9,2,0)
    worksheet.write(10,2,0)

arduino=serial.Serial('/dev/ttyACM0',9600)

#FECHA ACTUAL
now=time.strftime("%c")

workbook = xlsxwriter.Workbook('ingreso.xlsx')
worksheet = workbook.add_worksheet()
worksheet.set_column('A:A', 20)
writezero()
workbook.close()

leer=pd.read_excel("ingreso.xlsx")
codiguito=leer.iloc[:,0].values.tolist()
watch=leer.iloc[:,1].values.tolist()
datos=pd.read_csv('infochicos.csv',header=0)

#CODIGO
codigos=datos.iloc[:,0].values.tolist()
#NOMBRES
nombres=datos.iloc[:,1].values.tolist()
#CORREO
direccion=datos.iloc[:,2].values.tolist()
#UBICACION
acudientes=datos.iloc[:,4].values.tolist()
#CORREO
correos=datos.iloc[:,5].values.tolist()
#Codigos html
htmlcode=datos.iloc[:,6].values.tolist()
#EXCEL
posicion=datos.iloc[:,7].values.tolist()
#EXCEL
posifinal=datos.iloc[:,8].values.tolist()
#Codigos html
htmlcodevuelta=datos.iloc[:,9].values.tolist()
#Codigos html
htmlcodevueltados=datos.iloc[:,10].values.tolist()
{'a_code':codigos}

#MENSAJE 1
subject = "ATENCION, RUTA"
#MENSAJE2
subject2 = "MENSAJE, RUTA"
#DATOS CORREO
password = "santiagoyvalentina"
correoruta = "rutamigappoficial@gmail.com"
correosniños = correos
server = smtplib.SMTP('smtp.gmail.com: 587')
server.starttls()
server.login(correoruta, password)

app = Flask(__name__)
app.config['SECRET_KEY']='Thisisasecret'

class Loginform(FlaskForm):
    username=StringField('username',validators=[InputRequired('user name is requires'),Length(min=5,max=10,message='must  and 10 characteres')])
    password=PasswordField('password',validators=[InputRequired('password is requiered'),AnyOf(values=['rutauno','secret'],message='must  and 10 characteres')])

@app.route('/',methods=['GET','POST'])
def form():
    form=Loginform()
    if form.validate_on_submit():
        return render_template('botinicio.html')
    return render_template("login.html",form=form)

@app.route('/principal')
def principal():
    latlon=arduino.readline()
    latlon=latlon.decode()
    latlon=latlon.split(',')
    latitud=float(latlon[0])
    longitud=float(latlon[1])
    velocidad=float(latlon[2])
    ubicacion=[latitud,longitud,velocidad]
    print(ubicacion)
    if velocidad>80:
        print("Esta excediendo el limite de velocidad")
    return render_template('inicio.html')

@app.route('/<int:a_code>')
def getInformacion(a_code):
    for h in range(0,len(codigos)):
        codig=codigos[h]
        if a_code==codig:
            return render_template(htmlcode[h])

@app.route('/home')
def home():
    for b in range(0,len(codigos)):
        while 1:
            latlon=arduino.readline()
            latlon=latlon.decode()
            latlon=latlon.split(',')
            latitud=float(latlon[0])
            longitud=float(latlon[1])
            velocidad=float(latlon[2])
            ubicacion=[latitud,longitud,velocidad]
            print(ubicacion)
            ubicruta=ubicacion
            x=direccion[b]
            cerca=float(round(geodesic(x,ubicruta).meters))
            print(cerca)
            if velocidad>80:
                print("Esta excediendo el limite de velocidad")
                print("Esta excediendo el limite de velocidad")
                msg = "Esta excediendo el limite de velocidad"
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correoruta,message)
            if cerca<8000:
                print("Estimad@ "+str(acudientes[b])+", estoy a 5 minutos de recoger a "+str(nombres[b]))
                msg = "Estimad@ "+str(acudientes[b])+", estoy a 5 minutos de recoger a "+str(nombres[b])
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correosniños[b],message)
                return render_template(htmlcode[b])
    return render_template('inicio.html')

@app.route('/ingreso/<int:a_code>')
def getIngreso(a_code):
    print(a_code)
    doc = openpyxl.load_workbook('ingreso.xlsx')
    doc.get_sheet_names()
    [u'Sheet1']
    hoja = doc.get_sheet_by_name('Sheet1')
    for j in range(0,len(codigos)):
        code=codigos[j]
        if a_code==code:
            hoja[posicion[j]]=1
            hoja[posifinal[j]]=1
            doc.save('ingreso.xlsx')
            v=direccion[j]
            msg = "Estimad@ "+str(acudientes[j])+", "+str(nombres[j])+" acabo de ingresar a la ruta"
            message = 'Subject: {}\n{}'.format(subject,msg)
            server.sendmail(correoruta,correosniños[j],message)
            print("Esta recogiendo a ", nombres[j])
            confirmar=pd.read_excel("ingreso.xlsx")
            confir=confirmar.iloc[:,2].values.tolist()
            return render_template(htmlcode[j])

@app.route('/ausente/<int:a_code>')
def getAusente(a_code):
    doc = openpyxl.load_workbook('ingreso.xlsx')
    doc.get_sheet_names()
    [u'Sheet1']
    hoja = doc.get_sheet_by_name('Sheet1')
    for i in range(0,len(codigos)):
        cod=codigos[i]
        if a_code==cod:
            hoja[posifinal[i]]=1
            doc.save('ingreso.xlsx')
            print("No Ingreso ",nombres[i],". La fecha ",now)
            msg = "Estimad@ "+str(acudientes[i])+" la ruta espero 5 minutos, "+str(nombres[i])+" no ingreso a la ruta"
            message = 'Subject: {}\n{}'.format(subject,msg)
            server.sendmail(correoruta,correosniños[i],message)
            return render_template(htmlcode[i])

@app.route('/siguiente/<int:a_code>')
def getSiguiente(a_code):
    print(a_code)
    s=0
    for j in range(0,len(codigos)):
        code=codigos[j]
        while a_code==code:
            latlon=arduino.readline()
            latlon=latlon.decode()
            latlon=latlon.split(',')
            latitud=float(latlon[0])
            longitud=float(latlon[1])
            velocidad=float(latlon[2])
            ubicacion=[latitud,longitud,velocidad]
            print(ubicacion)
            ubicruta=ubicacion
            y=direccion[j]
            cerca2=float(round(geodesic(y,ubicruta).meters))
            print(cerca2)
            if velocidad>80:
                print("Esta excediendo el limite de velocidad")
                msg = "Esta excediendo el limite de velocidad"
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correoruta,message)
            if cerca2<8000:
                print("Estoy a 5 min de recoger a "+str(nombres[j]))
                msg = "Estimad@ "+str(acudientes[j])+" la ruta estara en 5 minutos, para recoger a "+str(nombres[j])
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correosniños[j],message)
                return render_template(htmlcode[j])

@app.route('/destino')
def destino():
    return render_template("destino.html")

@app.route('/finida')
def finida():
    doc = openpyxl.load_workbook('ingreso.xlsx')
    doc.get_sheet_names()
    [u'Sheet1']
    hoja = doc.get_sheet_by_name('Sheet1')
    look=pd.read_excel("ingreso.xlsx")
    centinela=look.iloc[:,2].values.tolist()
    for h in range(0,len(centinela)):
        if centinela[h]==1:
            print ("Estimad@ ", acudientes[h]," ", nombres[h], " ha llegado al colegio")
            msg = "Estimad@ "+str(acudientes[h])+", "+str(nombres[h])+" acabo de llegar a su destino"
            message = 'Subject: {}\n{}'.format(subject,msg)
            server.sendmail(correoruta,correosniños[h],message)
            workbook = xlsxwriter.Workbook('ingreso.xlsx')
            hoja[posicion[h]]=0
            hoja[posifinal[h]]=0
            doc.save('ingreso.xlsx')
    return render_template('botinicio.html')

@app.route('/principalvuelta')
def principalvuelta():
    return render_template('iniciovuelta.html')

@app.route('/regreso/<int:a_code>')
def getRegreso(a_code):
    for l in range(0,len(codigos)):
        codig=codigos[l]
        if a_code==codig:
            return render_template(htmlcodevuelta[l])

@app.route('/homevuelta')
def homevuelta():
    for m in range(0,len(codigos)):
        while 1:
            latlon=arduino.readline()
            latlon=latlon.decode()
            latlon=latlon.split(',')
            latitud=float(latlon[0])
            longitud=float(latlon[1])
            velocidad=float(latlon[2])
            ubicacion=[latitud,longitud,velocidad]
            print(ubicacion)
            ubicrutavuelta=ubicacion
            n=direccion[m]
            cerca=round(float(geodesic(n,ubicrutavuelta).meters))
            print(cerca)
            if velocidad>80:
                print("Esta excediendo el limite de velocidad")
                print("Esta excediendo el limite de velocidad")
                msg = "Esta excediendo el limite de velocidad"
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correoruta,message)
            if cerca<8000:
                print("estoy cerca a ",nombres[m])
                msg = "Estimad@ "+str(acudientes[m])+", "+str(nombres[m])+" estamos a 5 minutos de llegar a su hogar"
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correosniños[m],message)
                return render_template(htmlcodevueltados[m])
    return render_template('iniciovuelta.html')


@app.route('/vuelta/<int:a_code>')
def getVuelta(a_code):
    for k in range(0,len(codigos)):
        codig=codigos[k]
        while a_code==codig:
            latlon=arduino.readline()
            latlon=latlon.decode()
            latlon=latlon.split(',')
            latitud=float(latlon[0])
            longitud=float(latlon[1])
            velocidad=float(latlon[2])
            ubicacion=[latitud,longitud,velocidad]
            print(ubicacion)
            ubicruta=ubicacion
            z=direccion[k]
            cerca3=float(round(geodesic(z,ubicruta).meters))
            print(cerca3)
            if velocidad>80:
                print("Esta excediendo el limite de velocidad")
                msg = "Esta excediendo el limite de velocidad"
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correoruta,message)
            if cerca2<1000:
                print("Estoy a 5 min de recoger a "+str(nombres[k]))
                msg = "Estimad@ "+str(acudientes[k])+" la ruta estara en 5 minutos, para dejar a "+str(nombres[k])+"en la casa"
                message = 'Subject: {}\n{}'.format(subject,msg)
                server.sendmail(correoruta,correosniños[k],message)
            return render_template(htmlcodevuelta[k])

@app.route('/ingresovuelta/<int:a_code>')
def getIngresoVuelta(a_code):
    print(a_code)
    doc = openpyxl.load_workbook('ingreso.xlsx')
    doc.get_sheet_names()
    [u'Sheet1']
    hoja = doc.get_sheet_by_name('Sheet1')
    for o in range(0,len(codigos)):
        code=codigos[o]
        if a_code==code:
            hoja[posicion[o]]=1
            hoja[posifinal[o]]=1
            doc.save('ingreso.xlsx')
            v=direccion[o]
            msg = "Estimad@ "+str(acudientes[o])+", "+str(nombres[o])+" acabo de ingresar a la casa"
            message = 'Subject: {}\n{}'.format(subject,msg)
            server.sendmail(correoruta,correosniños[o],message)
            print("Esta dejando a ", nombres[o])
            confirmar=pd.read_excel("ingreso.xlsx")
            confir=confirmar.iloc[:,2].values.tolist()
            return render_template(htmlcodevueltados[o])

@app.route('/ausentevuelta/<int:a_code>')
def getAusenteVuelta(a_code):
    doc = openpyxl.load_workbook('ingreso.xlsx')
    doc.get_sheet_names()
    [u'Sheet1']
    hoja = doc.get_sheet_by_name('Sheet1')
    for p in range(0,len(codigos)):
        cod=codigos[p]
        if a_code==cod:
            hoja[posifinal[p]]=1
            doc.save('ingreso.xlsx')
            print("No Ingreso ",nombres[p],". La fecha ",now)
            msg = "Estimad@ "+str(acudientes[p])+" la ruta espero 5 minutos, "+str(nombres[p])+" no ingreso a la ruta"
            message = 'Subject: {}\n{}'.format(subject,msg)
            server.sendmail(correoruta,correosniños[p],message)
            return render_template(htmlcodevueltados[p])

@app.route('/finvuelta')
def finvuelta():
    doc = openpyxl.load_workbook('ingreso.xlsx')
    doc.get_sheet_names()
    [u'Sheet1']
    hoja = doc.get_sheet_by_name('Sheet1')
    look=pd.read_excel("ingreso.xlsx")
    centinela=look.iloc[:,2].values.tolist()
    for h in range(0,len(centinela)):
        if centinela[h]==1:
            print ("Estimad@ ", acudientes[h]," ", nombres[h], " ha llegado a su casa")
            msg = "Estimad@ "+str(acudientes[h])+", "+str(nombres[h])+" acabo de llegar a su casa"
            message = 'Subject: {}\n{}'.format(subject,msg)
            server.sendmail(correoruta,correosniños[h],message)
            workbook = xlsxwriter.Workbook('ingreso.xlsx')
            hoja[posicion[h]]=0
            hoja[posifinal[h]]=0
            doc.save('ingreso.xlsx')
    return render_template('botinicio.html')

if __name__ == "__main__":
    app.run(host='192.168.43.17',debug=True,port=4000)
